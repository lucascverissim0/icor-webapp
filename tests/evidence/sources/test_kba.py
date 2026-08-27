from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from icor.domain.evidence import MappingStatus, Measure, PublicationStatus, ReleaseManifest
from icor.evidence.serialization import sha256_file
from icor.evidence.sources.kba import KBAFZ10Loader
from icor.infrastructure.release_store import StoredRelease
from icor.infrastructure.sqlite_evidence_repository import SQLiteEvidenceRepository


def _stored_release(tmp_path: Path, *, total: int = 37, bad_header: bool = False) -> StoredRelease:
    artifact = tmp_path / "artifact.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FZ 10.1"
    sheet["B6"] = (
        "FZ 10.1 Neuzulassungen von Personenkraftwagen nach Marken und Modellreihen "
        "im Dezember 2024"
    )
    sheet["B9"] = "Fabrikat" if bad_header else "Marke"
    sheet["C9"] = "Modellreihe"
    sheet["D9"] = "Dezember  2024"
    sheet["E9"] = "Jan.-Dezember 2024"
    rows = [
        ("AUDI", "A1", 10),
        (None, "SONSTIGE", 2),
        ("AUDI ZUSAMMEN", "", 12),
        ("BMW", "X1", 20),
        ("BMW ZUSAMMEN", "", 20),
        ("SONSTIGE", "", 5),
        ("NEUZULASSUNGEN INSGESAMT", "", total),
    ]
    for row_number, (make, model, value) in enumerate(rows, start=10):
        sheet.cell(row_number, 2, make)
        sheet.cell(row_number, 3, model)
        sheet.cell(row_number, 5, value)
    sheet["B18"] = "© Kraftfahrt-Bundesamt, Flensburg"
    workbook.save(artifact)
    workbook.close()
    manifest = ReleaseManifest(
        release_id="kba-fz10-2024-12",
        source_id="kba-fz10",
        publisher="Kraftfahrt-Bundesamt",
        source_url="https://www.kba.de/SharedDocs/Downloads/DE/Statistik/Fahrzeuge/FZ10/fz10_2024_12.xlsx",
        retrieved_at=datetime(2026, 8, 27, tzinfo=UTC),
        published_at=datetime(2025, 1, 15, tzinfo=UTC),
        coverage_start=date(2024, 1, 1),
        coverage_end=date(2024, 12, 31),
        geography="DE",
        geography_version="DE-2024-v1",
        measure=Measure.NEW_REGISTRATIONS,
        unit="vehicles",
        publication_status=PublicationStatus.FINAL,
        dependency_group="de-zfzr-2024",
        terms_url="https://www.govdata.de/dl-de/by-2-0",
        permitted_local_use="Official open-data reuse with attribution",
        artifact_path="artifact.xlsx",
        artifact_bytes=artifact.stat().st_size,
        sha256=sha256_file(artifact),
        parser_name="kba_fz10_xlsx_v1",
        parser_version="v1",
        expected_schema="kba-fz10-2024-12-v1",
        raw_record_count=7,
        accepted_record_count=4,
        rejected_record_count=3,
        quarantined_record_count=0,
    )
    return StoredRelease(
        manifest.source_id, manifest.release_id, artifact, tmp_path / "manifest.json", manifest
    )


def test_loader_uses_annual_model_series_values_and_excludes_subtotals(tmp_path: Path) -> None:
    release = _stored_release(tmp_path)
    repository = SQLiteEvidenceRepository(tmp_path / "evidence.sqlite3", writable=True)
    repository.add_release(release.manifest)

    KBAFZ10Loader().load((release,), repository)

    rows = repository.list_observations()
    assert sorted((row.original_make, row.original_model, row.value) for row in rows) == [
        ("AUDI", "A1", Decimal("10")),
        ("AUDI", "SONSTIGE", Decimal("2")),
        ("BMW", "X1", Decimal("20")),
        ("SONSTIGE", "(not reported)", Decimal("5")),
    ]
    statuses = {(row.original_make, row.original_model): row.mapping_status for row in rows}
    assert statuses[("AUDI", "A1")] is MappingStatus.UNRESOLVED
    assert statuses[("AUDI", "SONSTIGE")] is MappingStatus.REJECTED
    assert statuses[("SONSTIGE", "(not reported)")] is MappingStatus.REJECTED


def test_loader_rejects_detail_total_that_does_not_match_published_total(tmp_path: Path) -> None:
    release = _stored_release(tmp_path, total=99)
    repository = SQLiteEvidenceRepository(tmp_path / "evidence.sqlite3", writable=True)
    repository.add_release(release.manifest)

    with pytest.raises(ValueError, match="published total"):
        KBAFZ10Loader().load((release,), repository)


def test_loader_rejects_header_drift(tmp_path: Path) -> None:
    release = _stored_release(tmp_path, bad_header=True)
    repository = SQLiteEvidenceRepository(tmp_path / "evidence.sqlite3", writable=True)
    repository.add_release(release.manifest)

    with pytest.raises(ValueError, match="header"):
        KBAFZ10Loader().load((release,), repository)
