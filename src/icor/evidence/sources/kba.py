"""KBA FZ10 December 2024 annual passenger-car registration adapter."""

from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from icor.domain.evidence import (
    EvidenceConfidence,
    MappingStatus,
    Measure,
    Observation,
    PeriodPrecision,
    PublicationStatus,
)
from icor.evidence.normalization import normalize_vehicle_label, stable_evidence_id
from icor.infrastructure.release_store import StoredRelease
from icor.infrastructure.sqlite_evidence_repository import SQLiteEvidenceRepository

PARSER_NAME = "kba_fz10_xlsx_v1"
SHEET_NAME = "FZ 10.1"
_TITLE = (
    "FZ 10.1 Neuzulassungen von Personenkraftwagen nach Marken und Modellreihen im Dezember 2024"
)
_HEADER = ("Marke", "Modellreihe", "Dezember  2024", "Jan.-Dezember 2024")


class KBAFZ10Loader:
    """Load KBA annual cumulative make/model-series counts without fuel double counting."""

    def load(
        self,
        releases: tuple[StoredRelease, ...],
        repository: SQLiteEvidenceRepository,
    ) -> None:
        for release in releases:
            self._load_release(release, repository)

    def _load_release(self, release: StoredRelease, repository: SQLiteEvidenceRepository) -> None:
        manifest = release.manifest
        if manifest.parser_name != PARSER_NAME:
            raise ValueError("KBA release parser name is unsupported")
        if manifest.publication_status is not PublicationStatus.FINAL:
            raise ValueError("KBA release must be final")
        if manifest.coverage_start != date(2024, 1, 1) or manifest.coverage_end != date(
            2024, 12, 31
        ):
            raise ValueError("KBA release must cover calendar year 2024")

        workbook = load_workbook(
            release.artifact_path, read_only=True, data_only=False, keep_links=False
        )
        try:
            if SHEET_NAME not in workbook.sheetnames:
                raise ValueError("KBA data worksheet is missing")
            sheet = workbook[SHEET_NAME]
            if sheet["B6"].value != _TITLE:
                raise ValueError("KBA workbook title is unsupported")
            if tuple(sheet.cell(9, column).value for column in range(2, 6)) != _HEADER:
                raise ValueError("KBA workbook header is unsupported")
            observations, raw_count, rejected_count, published_total = self._parse_rows(
                release, sheet
            )
        finally:
            workbook.close()

        accepted_count = len(observations)
        if (raw_count, accepted_count, rejected_count, 0) != (
            manifest.raw_record_count,
            manifest.accepted_record_count,
            manifest.rejected_record_count,
            manifest.quarantined_record_count,
        ):
            raise ValueError("KBA parser counts do not match manifest")
        detail_total = sum((row.value for row in observations), Decimal(0))
        if detail_total != published_total:
            raise ValueError("KBA detail registrations do not match published total")
        repository.add_observations(observations)

    def _parse_rows(
        self, release: StoredRelease, sheet: object
    ) -> tuple[list[Observation], int, int, Decimal]:
        observations: list[Observation] = []
        current_make: str | None = None
        raw_count = 0
        rejected_count = 0
        published_total: Decimal | None = None
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=10, min_col=2, max_col=5, values_only=False), start=10
        ):
            make_cell, model_cell, _, annual_cell = values
            make = _text(make_cell.value)
            model = _text(model_cell.value)
            if make is None and model is None and annual_cell.value is None:
                continue
            raw_count += 1
            if annual_cell.data_type == "f":
                raise ValueError("KBA annual value must not be a formula")
            annual = _vehicle_count(annual_cell.value)

            if make == "NEUZULASSUNGEN INSGESAMT":
                if published_total is not None:
                    raise ValueError("KBA published total appears more than once")
                published_total = Decimal(annual)
                rejected_count += 1
                break
            if make is not None and make.endswith(" ZUSAMMEN"):
                rejected_count += 1
                continue
            if make is not None:
                current_make = make
            if model is None and make != "SONSTIGE":
                raise ValueError("KBA detail row is missing a model series")
            if current_make is None:
                raise ValueError("KBA detail row is missing a make")

            source_model = model if model is not None else "(not reported)"
            mapping_status = (
                MappingStatus.REJECTED
                if current_make == "SONSTIGE" or source_model == "SONSTIGE" or model is None
                else MappingStatus.UNRESOLVED
            )
            notes = ["Read from the Jan.-Dezember 2024 annual cumulative total column."]
            flags: tuple[str, ...] = ()
            if model is None:
                notes.append("KBA published no model-series label for this unallocated row.")
                flags = ("source_model_missing",)
            elif model == "SONSTIGE":
                notes.append("KBA grouped low-volume or unlisted model series as SONSTIGE.")
                flags = ("source_model_aggregate",)
            observations.append(
                Observation(
                    observation_id=stable_evidence_id(
                        "obs-kba", release.release_id, str(row_number), current_make, source_model
                    ),
                    release_id=release.release_id,
                    original_row_locator=f"{SHEET_NAME}!E{row_number}",
                    geography="DE",
                    geography_version=release.manifest.geography_version,
                    period_start=date(2024, 1, 1),
                    period_end=date(2024, 12, 31),
                    period_precision=PeriodPrecision.YEAR,
                    measure=Measure.NEW_REGISTRATIONS,
                    value=Decimal(annual),
                    unit="vehicles",
                    publication_status=PublicationStatus.FINAL,
                    original_make=current_make,
                    original_model=source_model,
                    original_model_year=None,
                    original_type="KBA FZ10 model series",
                    source_make_identifier=current_make,
                    source_model_identifier=stable_evidence_id(
                        "kba-model", current_make, source_model
                    ),
                    normalized_make=normalize_vehicle_label(current_make),
                    normalized_model=normalize_vehicle_label(source_model),
                    normalized_model_year=None,
                    canonical_vehicle_id=None,
                    mapping_status=mapping_status,
                    transformation_notes=tuple(notes),
                    validation_flags=flags,
                    evidence_confidence=_confidence(mapping_status),
                )
            )
        if published_total is None:
            raise ValueError("KBA published total is missing")
        return observations, raw_count, rejected_count, published_total


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _vehicle_count(value: object) -> int:
    if value == "-":
        return 0
    if type(value) is not int or value < 0:
        raise ValueError("KBA annual registration count is invalid")
    return value


def _confidence(mapping_status: MappingStatus) -> EvidenceConfidence:
    identity = 0 if mapping_status is MappingStatus.REJECTED else 5
    return EvidenceConfidence(
        authority=25,
        publication_status=10,
        coverage=25,
        identity=identity,
        independent_agreement=10,
        reasons=(
            "Official finalized KBA Central Vehicle Register evidence.",
            "KBA model series is retained without an unreviewed cross-source alias.",
            "Agreement component is neutral until dependency-aware EEA overlap is evaluated.",
        ),
    )
